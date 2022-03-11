import json
import openpyxl
from collections import defaultdict

with open('data/input.json', 'r') as json_file:
	json_load = json.load(json_file)

new_dict = defaultdict(list)

comp=openpyxl.load_workbook("data/Software Company List v2.xlsx")

sh = comp.active
count=0
for i in range(1, sh.max_row+1):
    rt=False
    val=sh.cell(row=i, column=1).value.lower()
    val="".join(val.split(" "))
    val="".join(val.split("-"))
    val="".join(val.split("."))
    val="".join(val.split(","))
    for j in range(0,len(json_load)):
        ans=json_load[str(j)]["title"].lower()
        ans="".join(ans.split(" "))
        ans="".join(ans.split("-"))
        ans="".join(ans.split("."))
        ans="".join(ans.split(","))
        if ans.startswith(val):
            count+=1
            new_dict["0"].append(json_load[str(j)])
            rt=True
            break
    if rt==False:
        print(val)

print(count)
jdf=json.dumps(new_dict["0"])
# f = open("output.json", "w")
# f.write(jdf)
# f.close()